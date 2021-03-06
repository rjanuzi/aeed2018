from openpyxl import load_workbook
from openpyxl import Workbook

# In use cols
YEAR_COL = 0
PERIOD_COL = 1
FAMILY_COL = 2
QTD_COL = 3

inputFilePath = "C:\\Users\\Rafael\\Desktop\\aeed2018\\Projeto\\treated.xlsx"

def buildKey(year, period, family):
    return str(year) + "," + \
        str(period)  + "," + \
        str(family)

def getInput():
    print("Loading input")
    # Load Workbook. data_only = True force the formulas calculations
    wb = load_workbook(filename = inputFilePath, data_only = True) # Full input
    # Get the desired Sheet
    ws = wb['Sheet']

    # Create new dictionary
    inputs = dict()
    families = []

    # Start at row 2 to skip the table title
    for row in ws.iter_rows(min_row = 1, max_col = 5):
        family = row[FAMILY_COL].value.encode('utf-8')
        try:
            # Generate the map key
            temp_key = buildKey( \
                row[YEAR_COL].value, \
                row[PERIOD_COL].value, \
                family)

            # Adjust quantity to the map
            if temp_key in inputs:
                print("[ERROR]: input_loader.py - Loading - duplicated key: " + temp_key)
            else:
                inputs[temp_key] = row[QTD_COL].value

            if family not in families:
                families.append(family)

        except ValueError:
            print "Oops! Error interpreting the file"

    wb.close()
    print("Input Loaded: " + str(len(inputs)) + " entries with " + str(len(families)) + " families.")

    return inputs, families
