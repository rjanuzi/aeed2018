from openpyxl import load_workbook
from openpyxl import Workbook
from xlsWriter import saveEntriesDict
from xlsWriter import KnnResult

# In use cols
DATE_COL = 0
FAMILY_COL = 2
SALES_COL = 6

PERIOD_COUNT = 11*12
PERIOD_MIN = 1
PERIOD_MAX = PERIOD_COUNT+1
YEAR_MIN = 2010
YEAR_MAX = 2016 #2015 + 1

INPUT_FILE_PATH = "C:\\Users\\Rafael\\Desktop\\aeed2018\\Projeto\\Dataset\\Ztco0018 2010_2015.xlsx"
OUTPUT_FILE_PATH = "C:\\Users\\Rafael\\Desktop\\aeed2018\\Projeto\\treated.xlsx"

def calcPeriod(day, month):
        "This function calculate the period of year"
        month_part = calcMonthPart(day)

        if month_part <= 0:
                return -1
        return month_part + 6*month - 6

def calcPeriod_v2(day, month):
        "This function calculate the period of year"
        month_part = calcMonthPart_v2(day)

        if month_part <= 0:
                return -1
        return month_part + 11*month - 11

def calcMonthPart(day):
        "This function calculate the month part of the year's period"
        if(day <= 0):
                return -1
        # [1, 5)
        if day >= 1 and day < 5:
                return 1
        # [5, 10)
        if day >= 5 and day < 10:
                return 2
        # [10, 15)
        if day >= 10 and day < 15:
                return 3
        # [15, 20)
        if day >= 15 and day < 20:
                return 4
        # [20, 25)
        if day >= 20 and day < 25:
                return 5
        return 6

def calcMonthPart_v2(day):
        "This function calculate the month part of the year's period"
        if(day <= 0):
                return -1
        # [1, 3)
        if day >= 1 and day < 3:
                return 1
        # [3, 6)
        if day >= 3 and day < 6:
                return 2
        # [6, 9)
        if day >= 6 and day < 9:
                return 3
        # [9, 12)
        if day >= 9 and day < 12:
                return 4
        # [12, 15)
        if day >= 12 and day < 15:
                return 5
        # [15, 18)
        if day >= 15 and day < 18:
                return 6
        # [18, 21)
        if day >= 18 and day < 21:
                return 7
        # [21, 24)
        if day >= 21 and day < 24:
                return 8
        # [24, 27)
        if day >= 24 and day < 27:
                return 9
        # [27, 30)
        if day >= 27 and day < 30:
                return 10
        return 11

def buildKey(year, period, family):
    return str(year) + "," + \
        str(period)  + "," + \
        str(family)

print("Loading input")

# Load Workbook. data_only = True force the formulas calculations
wb = load_workbook(filename = INPUT_FILE_PATH, data_only = True) # Full input

# Get the desired Sheet
ws = wb['Plan1']

# Create new dictionary
sales_map = dict()
families = []

print("Processing input")

# Start at row 2 to skip the table title
for row in ws.iter_rows(min_row = 2, max_col = 13):
    tempFamilyName = row[FAMILY_COL].value.encode('utf-8') # Encode to solve a dataset problem
    try:
        # Generate the map key
        temp_key = buildKey(row[DATE_COL].value.year, \
                    str( calcPeriod_v2(row[DATE_COL].value.day, row[DATE_COL].value.month) ), \
                    tempFamilyName )

        # Adjust quantity to the map
        if temp_key in sales_map:
                sales_map[temp_key] = sales_map[temp_key] + row[SALES_COL].value
        else:
                sales_map[temp_key] = row[SALES_COL].value

        # Fill the families list for future use
        if tempFamilyName not in families:
            families.append(tempFamilyName)

    except ValueError:
        print "Oops! Error interpreting the file"

wb.close()

#Generate 0s for missing dots
print("Adding 0s for missing entries")

for i in range(len(families)):
    print("Adding 0s: " + str((float(i)/float(len(families)))*100.0) + "%")
    keysOfInterest = filter(lambda f: families[i] in f, sales_map.keys())
    for year in range(YEAR_MIN, YEAR_MAX):
        for period in range(PERIOD_MIN, PERIOD_MAX):
            tempKey = buildKey(year, period, families[i])
            if tempKey not in keysOfInterest:
                sales_map[tempKey] = 0
    i += 1

print("Saving treated.xlsx: Data: " + str(len(sales_map.keys())) + " with " + str(len(families)) + " families.")

saveEntriesDict(OUTPUT_FILE_PATH, sales_map)

print("Done")
