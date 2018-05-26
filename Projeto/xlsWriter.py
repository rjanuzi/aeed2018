from openpyxl import load_workbook
from openpyxl import Workbook

def saveEntriesDict(filePath, dictionary):
    res = Workbook()
    res_sheet = res.active

    for key, value in dictionary.items():
                splited_key = key.split(",")
                splited_key.append(value)
                res_sheet.append(splited_key)

    res.save(filePath)
    res.close()

class KnnResult:
    paramK = 0
    filterMaxMin = False
    useMedian = False
    averageError = 0.0
    maxError = 0.0
    minError = 0.0
    accuracy = 0.0

    def __repr__(self):
        return str(self.paramK) + "," + str(self.filterMaxMin) + "," + str(self.useMedian) + "," \
            + str(self.averageError) + "," + str(self.maxError) + "," + str(self.minError) + "," + str(self.accuracy)

def saveKnnResultList(knnResultList, filePath):
    res = Workbook()
    res_sheet = res.active

    res_sheet.append(["K", "Filter Max/Min", "Use Median", "Average Error", "Max Error", "Min Error", "Accuracy"])
    for result in knnResultList:
        res_sheet.append(str(result).split(","))

    res.save(filePath)
    res.close()
