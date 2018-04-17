from openpyxl import load_workbook

# Full input
#wb = load_workbook(filename = "input.xlsx")

# Reduced input
wb = load_workbook(filename = "input_min.xlsx")

ws = wb['Plan1']

for row in ws.iter_rows(min_row=1, max_col=13, max_row = 1003):
	print(row[1])
