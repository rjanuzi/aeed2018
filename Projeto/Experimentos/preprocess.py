from openpyxl import load_workbook
from openpyxl import Workbook

# In use cols
DATE_COL = 0
CLIENT_COL = 1
FAMILY_COL = 2
SALES_COL = 6

# Not in use cols
#SKU_COL = 3
#DESCRIPTION_COL = 4
#NF_COL = 5
#MATERIAL_GROUP_COL = 7
#CENTER_COL = 8
#BILING_TYPE_COL = 9
#PERIOD_COL = 10
#CLIENT_GROUP_COL = 11

def calcPeriod(day, month):
        "This function calculate the period of year"
        month_part = calcMonthPart(day)

        if month_part <= 0:
                return -1
        return month_part + 6*month - 6

def calcMonthPart(day):
        "This function calculate the month part of the year's period"
        if(day <= 0):
                return -1
        # [1, 5)
        if day >= 1 and day < 5:
                return 1
        # [5, 10)
        if day >= 5 and day < 10:
                return 2
        # [10, 15)
        if day >= 10 and day < 15:
                return 3
        # [15, 20)
        if day >= 15 and day < 20:
                return 4
        # [20, 25)
        if day >= 20 and day < 25:
                return 5
        return 6

# Load Workbook. data_only = True force the formulas calculations
wb = load_workbook(filename = "input.xlsx", data_only = True) # Full input
#wb = load_workbook(filename = "input_min.xlsx", data_only = True) # Reduced input

# Get the desired Sheet
ws = wb['Plan1']

# Create new dictionary
sales_map = dict()

# Start at row 2 to skip the table title
for row in ws.iter_rows(min_row = 2, max_col = 13):
        temp_key = str(row[DATE_COL].value.year) + "," + \
                str( calcPeriod(row[DATE_COL].value.day, row[DATE_COL].value.month) )  + "," + \
                str(row[CLIENT_COL].value) + "," + \
                str(row[FAMILY_COL].value)
        if temp_key in sales_map:
                sales_map[temp_key] = sales_map[temp_key] + row[SALES_COL].value
        else:
                sales_map[temp_key] = row[SALES_COL].value

res = Workbook()
res_sheet = res.active

for key, value in sales_map.items():
        splited_key = key.split(",")
        splited_key.append(value)
        res_sheet.append(splited_key)

res.save("result.xlsx")